import networkx as nx
import matplotlib.pyplot as plt
from openpyxl import load_workbook
import json

# Cargar el archivo Excel
wb = load_workbook("Ecuador_Distancias.xlsx", data_only=True)
sheet = wb["CUADRO DE DISTANCIAS"]

# Leer los nombres de las ciudades (columna B, desde la fila 4 en adelante)
ciudades = [sheet.cell(row=row, column=2).value for row in range(4, 44)]  # Ajusta el rango según tu archivo

# Crear un diccionario para almacenar las distancias
distancias = {}

# Leer las distancias desde el Excel
for row in range(4, 44):  # Ajusta el rango según las filas de tu archivo
    ciudad_origen = sheet.cell(row=row, column=2).value  # Columna B tiene los nombres de las ciudades
    distancias[ciudad_origen] = {
        ciudades[col - 3]: sheet.cell(row=row, column=col).value for col in range(3, 43)  # Columnas C en adelante
        if sheet.cell(row=row, column=col).value is not None
    }

# Crear el grafo
G = nx.Graph()

# Agregar nodos (ciudades) al grafo
G.add_nodes_from(ciudades)

# Agregar conexiones con sus distancias
for origen, destinos in distancias.items():
    for destino, distancia in destinos.items():
        if origen != destino and distancia > 0:  # Evitar distancias cero o conexiones consigo mismas
            G.add_edge(origen, destino, weight=distancia)

# Generar el Árbol de Expansión Mínima (MST)
mst = nx.minimum_spanning_tree(G)

# Dibujar el MST con mejoras en la presentación
plt.figure(figsize=(20, 15))  # Aumentar el tamaño del gráfico

# Usar un layout geográfico aproximado
posiciones = {
    "Ambato": (0, 3), "Azogues": (-4, -6), "Babahoyo": (-3, -2), "Cuenca": (-4, -8),
    "Esmeraldas": (-2, 6), "Guaranda": (-1, 3), "Guayaquil": (-5, -5), "Ibarra": (0, 6),
    "Latacunga": (0, 4), "Loja": (-5, -10), "Macas": (3, 0), "Machala": (-6, -7),
    "Nueva Loja": (3, 6), "Portoviejo": (-4, 2), "Orellana": (4, 3), "Puyo": (2, 2),
    "Quito": (0, 5), "Riobamba": (0, 2), "Tena": (2, 3), "Tulcán": (0, 7),
    "Zamora": (2, -8), "Aloag": (0, 4.5), "Santo Domingo": (-2, 4), "Baños": (1, 2),
    "Bahía de Caraquez": (-4, 4), "Baeza": (2, 5), "Rumichaca": (0, 8), "Macara": (-6, -9),
    "Huaquillas": (-7, -8), "Manta": (-5, 2), "Otavalo": (0, 6.5), "Salinas": (-6, -4),
    "San Lorenzo": (-1, 7), "Quevedo": (-3, 1), "Quininde": (-2, 5), "San Miguel": (1, -1),
    "Putumayo": (4, 7), "Morona": (3, -2), "Muisne": (-3, 5), "Pedernales": (-3, 6)
}

# Dibujar nodos con colores y tamaños personalizados
nx.draw_networkx_nodes(
    mst, posiciones, node_color='skyblue', node_size=1000, edgecolors='black', alpha=0.9
)

# Dibujar aristas con colores personalizados según las distancias
pesos = nx.get_edge_attributes(mst, 'weight')
aristas_colores = ['green' if peso < 100 else 'orange' if peso < 300 else 'red' for peso in pesos.values()]
nx.draw_networkx_edges(
    mst, posiciones, edge_color=aristas_colores, width=2, alpha=0.7
)

# Añadir etiquetas a los nodos con un tamaño de fuente más grande
nx.draw_networkx_labels(
    mst, posiciones, font_size=12, font_color='black', font_weight='bold'
)

# Añadir etiquetas a las aristas (distancias) con un tamaño de fuente más grande
nx.draw_networkx_edge_labels(
    mst, posiciones, edge_labels=pesos, font_size=10, label_pos=0.5, font_color='blue'
)

# Título del gráfico con un tamaño de fuente más grande
plt.title("Árbol de Expansión Mínima: Conexiones entre Ciudades del Ecuador", fontsize=20, fontweight='bold')

# Ocultar los ejes para una mejor presentación
plt.axis('off')

# Mostrar el gráfico
plt.show()

# Guardar las conexiones del MST en un archivo JSON
mst_conexiones = [(u, v, d['weight']) for (u, v, d) in mst.edges(data=True)]
with open("mst_conexiones_geograficas.json", "w", encoding="utf-8") as f:
    json.dump(mst_conexiones, f, ensure_ascii=False, indent=4)

print("Conexiones del MST guardadas en 'mst_conexiones_geograficas.json'.")